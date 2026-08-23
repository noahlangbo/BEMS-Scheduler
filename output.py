"""
output.py
=========
Exports the schedule to a formatted .xlsx and prints console summaries.

Sheets:
  1. Schedule        — week-by-week ambulance grid (Sun -> Sat)
  2. Campus Response — week-by-week campus responder grid (Sun -> Sat)
  3. Hour Summary    — per-person totals (ambulance + campus) vs requirements
  4. Warnings        — unfilled shifts, ALS without EVDT, night/weekend crews
                       without a driver, under-hours volunteers
  5. Strike List     — members whose submitted availability missed the minimums
"""

from __future__ import annotations

import csv
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import (
    CAMPUS_BLOCK_TIMES,
    SHIFT_TIMES,
    ShiftKey,
    Volunteer,
    is_big_weekend,
)

# ── Styling ──────────────────────────────────────────────────────────────────

C_HEADER_BG = "1F3864"
C_HEADER_FG = "FFFFFF"
C_DATE_BG = "2E5FA3"
C_DATE_FG = "FFFFFF"
C_EVDT_BG = "E2EFDA"
C_AUTH_BG = "FFF2CC"
C_EMT_BG = "FFFFFF"
C_BERT_BG = "D9E1F2"
C_WARN_BG = "FFE0E0"
C_ALT_ROW = "F8F8F8"
C_UNDER = "CC0000"

CERT_BG = {"EVDT": C_EVDT_BG, "Auth": C_AUTH_BG, "EMT": C_EMT_BG, "BERT": C_BERT_BG}

_thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _header_row(ws, row, values, widths=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _font(bold=True, color=C_HEADER_FG)
        c.fill = _fill(C_HEADER_BG)
        c.alignment = _align(h="center")
        c.border = BORDER
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w


def _week_start_sunday(d: date) -> date:
    return d - timedelta(days=(d.weekday() + 1) % 7)


DAY_NAMES = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]


def _weekly_grid(ws, keys_to_people, sections, slot_rows, cell_for):
    """
    Shared week-by-week grid builder.
      sections:  [(section_key, time_label)] — one band of rows per section
      slot_rows: [slot_label, ...] — rows inside each band
      cell_for:  (people, slot_index) -> (text, fill_hex)
    """
    widths = [10, 12] + [18] * 7
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    if not keys_to_people:
        return

    all_dates = sorted({d for (d, _) in keys_to_people})
    current, end = _week_start_sunday(all_dates[0]), all_dates[-1]

    row = 1
    while current <= end:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = ws.cell(row=row, column=1, value=f"Week of {current.isoformat()} (Sun–Sat)")
        c.font = _font(bold=True, color=C_DATE_FG)
        c.fill = _fill(C_DATE_BG)
        c.alignment = _align(h="left")
        c.border = BORDER
        row += 1

        headers = ["", "Slot"] + [
            f"{DAY_NAMES[i]} {(current + timedelta(days=i)).strftime('%m/%d')}" for i in range(7)
        ]
        _header_row(ws, row, headers)
        row += 1

        for section_key, time_label in sections:
            band_start = row
            for slot_idx, slot_label in enumerate(slot_rows):
                c = ws.cell(row=row, column=2, value=slot_label)
                c.border = BORDER
                c.alignment = _align(h="center")
                c.font = _font(bold=True)
                for i in range(7):
                    d = current + timedelta(days=i)
                    people = keys_to_people.get((d, section_key))
                    text, fill_hex = ("", C_EMT_BG)
                    if people is not None:
                        text, fill_hex = cell_for(people, slot_idx)
                    cell = ws.cell(row=row, column=3 + i, value=text)
                    cell.fill = _fill(fill_hex)
                    cell.border = BORDER
                    cell.alignment = _align(h="left", wrap=True)
                    cell.font = _font(size=9, bold=bool(text))
                row += 1

            ws.merge_cells(start_row=band_start, start_column=1, end_row=row - 1, end_column=1)
            c = ws.cell(row=band_start, column=1, value=f"{section_key}\n{time_label}")
            c.alignment = _align(h="center", wrap=True)
            c.font = _font(bold=True)
            for r in range(band_start, row):
                ws.cell(row=r, column=1).border = BORDER
                ws.cell(row=r, column=1).fill = _fill(C_ALT_ROW)
            row += 1  # spacer between bands

        row += 1  # spacer between weeks
        current += timedelta(days=7)


# ── Sheet 1: Ambulance schedule ──────────────────────────────────────────────

def _build_schedule_sheet(ws, assignments):
    ws.title = "Schedule"
    ws.freeze_panes = "C3"

    sections = [(s, f"{SHIFT_TIMES[s][0]}-{SHIFT_TIMES[s][1]}") for s in ("DAY", "AM", "PM", "NIGHT")]

    def cell_for(people, slot_idx):
        # Rows: EVDTs, Auths, then EMT-only in order.
        evdts = [v for v in people if v.is_evdt]
        auths = [v for v in people if v.certification == "Auth"]
        emts = [v for v in people if not v.is_driver]
        if slot_idx == 0:
            return ("\n".join(v.full_name for v in evdts), C_EVDT_BG if evdts else C_EMT_BG)
        if slot_idx == 1:
            return ("\n".join(v.full_name for v in auths), C_AUTH_BG if auths else C_EMT_BG)
        if slot_idx == 2:
            return (emts[0].full_name if emts else "", C_EMT_BG)
        # Last row stacks any remaining EMTs so no one falls off the grid.
        return ("\n".join(v.full_name for v in emts[1:]), C_EMT_BG)

    _weekly_grid(ws, assignments, sections, ["EVDT", "Auth", "EMT1", "EMT2"], cell_for)


# ── Sheet 2: Campus response ─────────────────────────────────────────────────

def _build_campus_sheet(ws, campus_assignments, responders_per_block: int):
    ws.title = "Campus Response"
    ws.freeze_panes = "C3"

    sections = [(b, CAMPUS_BLOCK_TIMES[b]) for b in ("A", "B", "C", "D")]
    slot_rows = [f"Responder{i + 1}" for i in range(responders_per_block)]

    def cell_for(people, slot_idx):
        if slot_idx < len(people):
            p = people[slot_idx]
            return (p.full_name, CERT_BG.get(p.certification, C_BERT_BG))
        return ("", C_EMT_BG)

    _weekly_grid(ws, campus_assignments, sections, slot_rows, cell_for)


# ── Sheet 3: Hour summary ────────────────────────────────────────────────────

def _fmt_keys(keys) -> str:
    return "; ".join(f"{d.month}/{d.day} {s}" for d, s in sorted(keys))


def _build_summary_sheet(ws, people, ambulance_required, campus_emt_required, campus_bert_required):
    ws.title = "Hour Summary"
    ws.freeze_panes = "A2"
    headers = ["Name", "Email", "Role", "Certification", "Ambulance Hours", "Ambulance Shifts",
               "Campus Hours", "Campus Blocks", "Ambulance Status", "Campus Status"]
    widths = [28, 32, 8, 13, 14, 44, 12, 30, 16, 16]
    _header_row(ws, 1, headers, widths)

    def sort_key(p):
        return (-getattr(p, "assigned_hours", 0), -p.campus_assigned_hours, p.full_name)

    for i, p in enumerate(sorted(people, key=sort_key), 2):
        is_bert = not isinstance(p, Volunteer)
        amb = 0 if is_bert else p.assigned_hours
        campus = p.campus_assigned_hours
        campus_required = campus_bert_required if is_bert else campus_emt_required
        amb_under = (not is_bert) and amb < ambulance_required
        campus_under = campus < campus_required
        amb_status = "—" if is_bert else (f"⚠ {amb}/{ambulance_required}h" if amb_under else "OK")
        campus_status = f"⚠ {campus}/{campus_required}h" if campus_under else "OK"
        vals = [
            p.full_name, p.email, "BERT" if is_bert else "AMB", p.certification,
            "—" if is_bert else amb,
            "—" if is_bert else _fmt_keys(p.assigned),
            campus, _fmt_keys(p.campus_assigned),
            amb_status, campus_status,
        ]
        bg = C_ALT_ROW if i % 2 == 0 else C_EMT_BG
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = _fill(bg)
            under = (col == 9 and amb_under) or (col == 10 and campus_under)
            c.font = _font(bold=under, color=C_UNDER if under else "000000")
            c.alignment = _align(h="center" if col in (3, 4, 5, 7, 9, 10) else "left",
                                 wrap=col in (6, 8))
            c.border = BORDER


# ── Sheet 4: Warnings ────────────────────────────────────────────────────────

def collect_warnings(
    assignments,
    als_shifts,
    volunteers,
    ambulance_required,
    campus_assignments=None,
    responders_per_block: int = 2,
) -> list[tuple]:
    """[(type, date, shift, details)] for everything worth a human look."""
    issues = []
    for key in sorted(assignments):
        d, s = key
        people = assignments[key]
        if not people:
            issues.append(("UNFILLED SHIFT", d, s, "No volunteers assigned"))
            continue
        if key in als_shifts and not any(v.is_evdt for v in people):
            issues.append(("ALS — NO EVDT", d, s,
                           f"Assigned: {', '.join(v.full_name for v in people)}"))
        needs_driver = s == "NIGHT" or is_big_weekend(d, s)
        if needs_driver and not any(v.is_driver for v in people):
            issues.append(("NO DRIVER", d, s, "No EVDT or Auth on crew"))
    for v in sorted(volunteers, key=lambda v: v.full_name):
        if v.assigned_hours < ambulance_required:
            issues.append(("UNDER HOURS", None, "",
                           f"{v.full_name}: {v.assigned_hours}/{ambulance_required}h ambulance"))
    if campus_assignments is not None:
        for (d, block), people in sorted(campus_assignments.items()):
            if not any(getattr(p, "is_driver", False) for p in people):
                issues.append(("CAMPUS — OPEN S1", d, block,
                               "No driver-eligible responder assigned; block left open"))
            elif len(people) < responders_per_block:
                issues.append(("CAMPUS — OPEN SEAT", d, block,
                               f"{len(people)}/{responders_per_block} responders assigned"))
    return issues


def _build_warnings_sheet(ws, issues):
    ws.title = "Warnings"
    headers = ["Type", "Date", "Day", "Shift", "Details"]
    widths = [22, 13, 12, 8, 50]
    _header_row(ws, 1, headers, widths)

    if not issues:
        c = ws.cell(row=2, column=1, value="No warnings — all shifts adequately staffed.")
        c.font = _font(bold=True, color="1E7E34")
        return

    for i, (type_, d, shift_type, detail) in enumerate(issues, 2):
        bg = "FFF0F0" if i % 2 == 0 else C_WARN_BG
        vals = [type_, d.isoformat() if d else "", d.strftime("%A") if d else "", shift_type, detail]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = _fill(bg)
            c.font = _font(bold=(col == 1), color=C_UNDER if col == 1 else "000000")
            c.alignment = _align(h="center" if col in (2, 3, 4) else "left")
            c.border = BORDER


# ── Sheet 5: Strike list ─────────────────────────────────────────────────────

def _build_strike_list_sheet(ws, violations):
    ws.title = "Strike List"
    ws.freeze_panes = "A2"
    headers = ["Name", "Email", "Certification", "Missing Requirements"]
    widths = [28, 32, 14, 50]
    _header_row(ws, 1, headers, widths)

    if not violations:
        c = ws.cell(row=2, column=1, value="✓ Everyone met the minimum availability requirements.")
        c.font = _font(bold=True, color="1E7E34")
        return

    for i, item in enumerate(violations, 2):
        v = item["volunteer"]
        bg = C_ALT_ROW if i % 2 == 0 else C_EMT_BG
        vals = [v.full_name, v.email, v.certification, ", ".join(item["missing"])]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = _fill(bg)
            c.font = _font(bold=(col == 1))
            c.alignment = _align(h="center" if col == 3 else "left", wrap=(col == 4))
            c.border = BORDER


# ── Entry points ─────────────────────────────────────────────────────────────

def export_schedule_xlsx(
    assignments: dict[ShiftKey, list],
    campus_assignments: dict[ShiftKey, list],
    people: list,
    output_path: str,
    als_shifts: set[ShiftKey],
    violations: list[dict],
    ambulance_required: int,
    campus_emt_required: int,
    campus_bert_required: int,
    responders_per_block: int = 2,
) -> str:
    if not output_path.endswith(".xlsx"):
        output_path = output_path.rsplit(".", 1)[0] + ".xlsx"

    volunteers = [p for p in people if isinstance(p, Volunteer)]
    issues = collect_warnings(
        assignments, als_shifts, volunteers, ambulance_required,
        campus_assignments, responders_per_block,
    )

    wb = Workbook()
    _build_schedule_sheet(wb.active, assignments)
    _build_campus_sheet(wb.create_sheet(), campus_assignments, responders_per_block)
    _build_summary_sheet(wb.create_sheet(), people, ambulance_required,
                         campus_emt_required, campus_bert_required)
    _build_warnings_sheet(wb.create_sheet(), issues)
    _build_strike_list_sheet(wb.create_sheet(), violations)
    wb.save(output_path)
    print(f"  Schedule exported -> {output_path}")
    return output_path


def print_summary(
    assignments: dict[ShiftKey, list],
    campus_assignments: dict[ShiftKey, list],
    volunteers: list[Volunteer],
    bert_members: list,
    als_shifts: set[ShiftKey],
    ambulance_required: int,
    campus_emt_required: int,
    campus_bert_required: int,
    responders_per_block: int = 2,
) -> None:
    total = len(assignments)
    unfilled = sum(1 for p in assignments.values() if not p)
    filled_slots = sum(len(p) for p in assignments.values())
    als_no_evdt = sum(
        1 for k, p in assignments.items() if k in als_shifts and not any(v.is_evdt for v in p)
    )
    no_driver = sum(
        1 for (d, s), p in assignments.items()
        if (s == "NIGHT" or is_big_weekend(d, s)) and not any(v.is_driver for v in p)
    )
    under = sum(1 for v in volunteers if v.assigned_hours < ambulance_required)

    c_total = len(campus_assignments)
    c_unfilled = sum(1 for p in campus_assignments.values() if not p)
    c_full = sum(1 for p in campus_assignments.values() if len(p) >= responders_per_block)
    emt_campus_under = sum(1 for v in volunteers if v.campus_assigned_hours < campus_emt_required)
    bert_under = sum(1 for b in bert_members if b.campus_assigned_hours < campus_bert_required)

    print("\n" + "=" * 55)
    print("SCHEDULE SUMMARY")
    print("=" * 55)
    print(f"  Ambulance shifts:            {total}")
    print(f"    Unfilled:                  {unfilled}" + (" ⚠" if unfilled else " ✓"))
    print(f"    Filled crew slots:         {filled_slots}")
    print(f"    ALS without EVDT:          {als_no_evdt}" + (" ⚠" if als_no_evdt else " ✓"))
    print(f"    Night/weekend w/o driver:  {no_driver}" + (" ⚠" if no_driver else " ✓"))
    print(f"    EMTs under {ambulance_required}h:             {under}" + (" ⚠" if under else " ✓"))
    print(f"  Campus blocks:               {c_total}")
    print(f"    Unfilled:                  {c_unfilled}" + (" ⚠" if c_unfilled else " ✓"))
    print(f"    Fully staffed (={responders_per_block}):        {c_full}")
    print(f"    EMTs under {campus_emt_required}h campus:       {emt_campus_under}" + (" ⚠" if emt_campus_under else " ✓"))
    print(f"    BERT under {campus_bert_required}h campus:       {bert_under}" + (" ⚠" if bert_under else " ✓"))
    print()


def print_warnings(issues: list[tuple]) -> None:
    print("=" * 55)
    print("WARNINGS")
    print("=" * 55)
    if not issues:
        print("  No warnings.")
    for type_, d, shift_type, detail in issues:
        where = f"{d.isoformat()} ({d.strftime('%a')}) {shift_type}" if d else detail
        extra = f" — {detail}" if d and detail else ""
        print(f"  ⚠  {type_}: {where}{extra}")
    print()


# ── Master Schedule CSV ─────────────────────────────────────────────────────

MASTER_SCHEDULE_HEADER = [
    "Block", "ShiftID", "Date", "Shift", "Vehicle", "Seat", "Requires", "Assigned/Name",
]


def _date_for_master_schedule(d: date) -> str:
    """Portable M/D/YY formatting (strftime %-m is not portable to Windows)."""
    return f"{d.month}/{d.day}/{d:%y}"


def _day_number(d: date, block_start: date, daynum_start: int) -> str:
    return f"{daynum_start + (d - block_start).days:04d}"


def _ambulance_master_rows(assignments, block_start, block, daynum_start, vehicle):
    rows = []
    for (d, shift), people in sorted(assignments.items()):
        if not people:
            continue
        daynum = _day_number(d, block_start, daynum_start)
        # Preserve every assignment while putting the preferred driver in the
        # driver's row.  The previous draft dropped additional driver-eligible
        # crew members entirely.
        primary_driver = next((p for p in people if getattr(p, "is_driver", False)), None)
        ordered = ([primary_driver] if primary_driver else []) + [p for p in people if p is not primary_driver]
        for i, person in enumerate(ordered):
            if i == 0 and primary_driver is not None:
                seat, requires, suffix = "Driver", "EVDT", "EVDT"
            else:
                crew_number = i + 1 if primary_driver is not None else i + 2
                seat, requires, suffix = f"C{crew_number}", "CREW", f"C{crew_number}"
            shift_id = f"{block}-{daynum}-{shift}-{vehicle}-{suffix}"
            rows.append([block, shift_id, _date_for_master_schedule(d), shift,
                         vehicle, seat, requires, person.full_name])
    return rows


def _campus_master_rows(campus_assignments, block_start, block, daynum_start):
    rows = []
    for (d, campus_block), people in sorted(campus_assignments.items()):
        daynum = _day_number(d, block_start, daynum_start)
        ordered = sorted(people, key=lambda p: (not getattr(p, "is_driver", False), p.full_name))
        for i, person in enumerate(ordered, start=1):
            seat = f"S{i}"
            shift_id = f"{block}-{daynum}-{campus_block}-CR-{seat}"
            rows.append([block, shift_id, _date_for_master_schedule(d), campus_block,
                         "CR", seat, "AUTH" if seat == "S1" else "CREW", person.full_name])
    return rows


def export_master_schedule_csv(
    assignments,
    campus_assignments,
    block_start: date,
    output_path: str = "master_schedule.csv",
    block: str = "F26B1",
    daynum_start: int = 810,
    vehicle: str = "R1",
) -> str:
    """Write flat A:H rows that can be pasted into Master Schedule row 2."""
    rows = _ambulance_master_rows(assignments, block_start, block, daynum_start, vehicle)
    rows += _campus_master_rows(campus_assignments, block_start, block, daynum_start)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(MASTER_SCHEDULE_HEADER)
        writer.writerows(rows)
    print(f"  Master Schedule CSV exported -> {output_path} ({len(rows)} rows)")
    return output_path
