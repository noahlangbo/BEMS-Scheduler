"""Regression tests for the unchanged Fall 2026 form, including appended columns.

The fixture preserves relevant live headers/order, not private response data.
"""
import csv
import io
import json
import tempfile
import unittest
from contextlib import redirect_stdout
from datetime import date
from pathlib import Path
from unittest.mock import patch

from models import Volunteer
from parse_form import _build_column_maps, load_all_responses, normalise_driver
from validate import AvailabilityRequirements, check_ambulance_requirements, check_bert_requirements

START, END = date(2026, 9, 8), date(2026, 9, 20)
HEADERS = json.loads(Path(__file__).with_name("shopping_headers.json").read_text())


def response(role="Ambulance EMT (EMT only & EMT/ERT dual-role)", **answers):
    row = [""] * len(HEADERS)
    row[0:2] = ["8/31/2026 17:03:57", "fixture@example.invalid"]
    row[8] = role
    if role == "BERT Member Only":
        row[31:33] = ["Responder", "Test"]
        row[43] = "N/A"
    else:
        row[9:12] = ["EMT", "Test", "EVDT - Rescue 1/Utility 1"]
        row[30] = "N/A"
    for index, value in answers.items():
        row[int(index)] = value
    return row


def parse(rows, headers=HEADERS, start=START, end=END):
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "responses.csv"
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            csv.writer(f).writerows([headers] + rows)
        return load_all_responses(str(path), start, end)


class ShoppingParserTests(unittest.TestCase):
    def test_live_header_layout_recognizes_every_availability_column(self):
        maps = _build_column_maps(HEADERS, START, END)
        self.assertEqual(len(maps["emt_week"]), 9)
        self.assertEqual(len(maps["emt_night"]), 6)
        self.assertEqual(len(maps["emt_weekend"]), 4)
        self.assertEqual(len(maps["bert"]), 9)
        self.assertEqual(maps["emt_night"][44], date(2026, 9, 13))
        self.assertEqual(maps["emt_night"][45], END)

    def test_original_test_selections_are_preserved(self):
        v, b = parse([response(**{"13": "AM, PM, NIGHT", "15": "PM", "16": "NIGHT",
                                  "20": "PM", "22": "PM, NIGHT", "25": "NIGHT",
                                  "26": "NIGHT", "27": "DAY", "28": "DAY"})])
        self.assertEqual(len(v[0].available), 12)
        self.assertEqual(len(v[0].campus_available), 10)
        self.assertEqual(v[0].certification, "EVDT")
        self.assertEqual(b, [])

    def test_sunday_nights_appended_after_cr_are_ambulance_only(self):
        v, _ = parse([response(**{"44": "NIGHT", "45": "NIGHT (1900-0700)"})])
        self.assertEqual(v[0].available, {(date(2026, 9, 13), "NIGHT"), (END, "NIGHT")})
        self.assertFalse(v[0].campus_available)

    def test_bert_branch_uses_own_names_and_a_b_c_d(self):
        _, b = parse([response("BERT Member Only", **{"33": "A (0700-1000), C (1300-1600)",
                                                     "41": "Block B, Block D", "44": "NIGHT"})])
        self.assertEqual(b[0].full_name, "Test Responder")
        self.assertEqual(len(b[0].campus_available), 4)
        self.assertEqual(check_bert_requirements(b, AvailabilityRequirements()), [])

    def test_no_availability_is_allowed_when_columns_exist(self):
        v, _ = parse([response()])
        self.assertFalse(v[0].available)

    def test_duplicate_friday_night_is_counted_once(self):
        v, _ = parse([response(**{"15": "NIGHT", "16": "NIGHT"})])
        self.assertEqual(v[0].available, {(date(2026, 9, 11), "NIGHT")})

    def test_blank_not_available_and_invalid_day_tokens(self):
        v, _ = parse([response(**{"12": "NOT AVAILABLE", "13": "AM, NOT AVAILABLE",
                                  "18": "AM, DAY", "44": "PM, NIGHT"})])
        self.assertEqual(v[0].available, {(date(2026, 9, 9), "AM"),
                                         (date(2026, 9, 12), "DAY"), (date(2026, 9, 13), "NIGHT")})

    def test_latest_submission_wins(self):
        old = response(**{"12": "AM"})
        new = response(**{"13": "PM"})
        new[0] = "9/1/2026 18:00:00"
        v, _ = parse([new, old])
        self.assertEqual(v[0].available, {(date(2026, 9, 9), "PM")})

    def test_driver_options(self):
        self.assertEqual(normalise_driver("EVDT - Rescue 1/Utility 1"), "EVDT")
        self.assertEqual(normalise_driver("Authorized - Utility 1"), "Auth")
        self.assertEqual(normalise_driver("Not a driver"), "EMT")

    def test_blackouts_still_apply(self):
        v, _ = parse([response(**{"13": "AM, PM", "30": "9/9 AM"})])
        self.assertEqual(v[0].available, {(date(2026, 9, 9), "PM")})

    def test_old_august_config_stops_instead_of_empty_schedule(self):
        with self.assertRaisesRegex(ValueError, "No recognized Ambulance"):
            parse([response()], start=date(2026, 8, 10), end=date(2026, 9, 3))

    def test_unknown_availability_header_stops(self):
        headers = HEADERS.copy()
        headers[12] = "Changed question [Tue 9/8]"
        with self.assertRaisesRegex(ValueError, "Unrecognized availability"):
            parse([response()], headers)

    def test_missing_email_header_stops(self):
        headers = HEADERS.copy()
        headers[1] = "Something else"
        with self.assertRaisesRegex(ValueError, "Missing required form column"):
            parse([response()], headers)

    def test_unknown_role_stops(self):
        with self.assertRaisesRegex(ValueError, "Unrecognized member role"):
            parse([response("Unexpected role")])

    def test_missing_cr_columns_stops_for_bert_response(self):
        headers = HEADERS.copy()
        for i in range(33, 42):
            headers[i] = "Unused"
        with self.assertRaisesRegex(ValueError, "No recognized Campus Response"):
            parse([response("BERT Member Only")], headers)


class ShoppingRequirementTests(unittest.TestCase):
    def setUp(self):
        self.cfg = json.loads((Path(__file__).parent.parent / "config.json").read_text())
        self.reqs = AvailabilityRequirements.from_config(self.cfg)
        self.slots = {(date(2026, 9, 8), "AM"), (date(2026, 9, 8), "PM"),
                      (date(2026, 9, 13), "NIGHT"), (date(2026, 9, 12), "DAY"),
                      (date(2026, 9, 11), "NIGHT")}

    def check(self, slots):
        return check_ambulance_requirements([Volunteer("Test", "EMT", "test@example.invalid",
                                                       "EMT", available=slots)], self.reqs)

    def test_each_of_five_categories_is_required(self):
        self.assertFalse(self.check(self.slots))
        for slot in self.slots:
            with self.subTest(slot=slot):
                self.assertEqual(len(self.check(self.slots - {slot})[0]["missing"]), 1)

    def test_sunday_night_does_not_satisfy_weekend_night(self):
        missing = self.check(self.slots - {(date(2026, 9, 11), "NIGHT")})[0]["missing"]
        self.assertEqual(missing, ["Weekend NIGHT (0/1)"])

    def test_friday_night_does_not_satisfy_weekday_night(self):
        missing = self.check(self.slots - {(date(2026, 9, 13), "NIGHT")})[0]["missing"]
        self.assertEqual(missing, ["Weekday NIGHT (0/1)"])

    def test_two_pm_and_saturday_night_no_longer_pass(self):
        slots = {(date(2026, 9, 8), "PM"), (date(2026, 9, 9), "PM"), (date(2026, 9, 12), "NIGHT")}
        self.assertEqual(len(self.check(slots)[0]["missing"]), 3)

    def test_legacy_thresholds_remain_supported(self):
        v = Volunteer("Test", "Legacy", "test@example.invalid", "EMT", available={
            (date(2026, 9, 8), "PM"), (date(2026, 9, 9), "PM"), (date(2026, 9, 12), "NIGHT")})
        self.assertFalse(check_ambulance_requirements([v], AvailabilityRequirements.from_config({})))

    def test_block_config_matches_form(self):
        self.assertEqual(self.cfg["block_start"], START.isoformat())
        self.assertEqual(self.cfg["block_end"], END.isoformat())
        self.assertEqual(self.cfg["hours"], {"ambulance_emt": 12, "campus_emt": 3, "campus_bert": 3})


class ShoppingPipelineTests(unittest.TestCase):
    def test_form_to_workbook_and_master_csv(self):
        from main import main
        from openpyxl import load_workbook

        cfg = json.loads((Path(__file__).parent.parent / "config.json").read_text())
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            cfg["form_csv"] = str(root / "responses.csv")
            cfg["output_xlsx"] = str(root / "output.xlsx")
            cfg["master_schedule_export"]["path"] = str(root / "master.csv")
            cfg["solver_time_limit_s"] = 2
            with open(cfg["form_csv"], "w", newline="") as f:
                csv.writer(f).writerows([HEADERS, response(**{
                    "12": "AM, PM, NIGHT", "16": "NIGHT", "18": "DAY",
                    "44": "NIGHT", "45": "NIGHT"}),
                    response("BERT Member Only", **{"33": "A, B, C, D"})])
            config_path = root / "config.json"
            config_path.write_text(json.dumps(cfg))
            with patch("sys.argv", ["main.py", str(config_path)]), redirect_stdout(io.StringIO()):
                main()
            workbook = load_workbook(cfg["output_xlsx"])
            self.assertTrue(workbook.sheetnames)
            workbook.close()
            with open(cfg["master_schedule_export"]["path"], newline="") as f:
                rows = list(csv.DictReader(f))
            self.assertTrue(rows)
            self.assertEqual({r["Block"] for r in rows}, {"F26SHOP"})
            self.assertTrue(any(r["ShiftID"].startswith("F26SHOP-0908-") for r in rows))


if __name__ == "__main__":
    unittest.main()
